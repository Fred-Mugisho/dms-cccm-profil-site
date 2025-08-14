from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status

from data_import.forms import MouvementDeplaceSiteUniqueForm
from .models import *
from django.http import HttpResponse
from django.db import transaction
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from utils.functions import *
from .utils import DataImportService

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import authenticate, login
from .serializers import MouvementDeplaceSiteUniqueSerializer
import requests
from django.db.models import Q

@api_view(['POST'])
def import_data(request):
    """Import des données CCCM depuis un fichier Excel - avec sécurité transactionnelle"""
    try:
        service = DataImportService()
        file_data = request.FILES.get('file_data')

        if not file_data:
            return Response({"message": "Le fichier de données est requis"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file_data, data_only=True, read_only=True)
        except Exception as e:
            return Response({"message": f"Erreur lors du chargement du fichier Excel: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():  # ⛔ Début du bloc atomique
            # Nettoyage des anciens imports
            TemporalMouvementDeplace.objects.all().delete()
            MouvementDeplace.objects.all().delete()
            SiteDeplace.objects.all().delete()
            logging.info("Anciennes données supprimées.")
            
            logging.info("Lecture des feuilles Excel...")

            # Lecture et traitement des feuilles Excel
            for sheet_name, default_date in service.SHEETS_CONFIG:
                if sheet_name not in wb.sheetnames:
                    logging.warning(f"Feuille {sheet_name} non trouvée")
                    continue

                sheet = wb[sheet_name]

                service.process_sheet_data_v3(sheet, sheet_name, default_date)
                
            variation_result = service.generate_variation_data()
            logging.info(f"Variations: {variation_result}")

            logging.info("Importation terminée.")

        return Response(service.statistiques(), status=status.HTTP_200_OK)
    except Exception as e:
        logging.error(f"Erreur lors de l'importation: {e}")
        return Response({"message": f"Erreur lors de l'importation: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def import_data_site(request):
    """Import des données CCCM depuis un fichier Excel - avec sécurité transactionnelle"""
    try:
        service = DataImportService()
        
        API_ENDPOINT = "http://dmscccm.wnhelp.org/api/data-import/mouvements_deplaces/"

        # --- Étape 1 : Récupération des données ---
        try:
            response = requests.get(API_ENDPOINT, timeout=60)
            response.raise_for_status()
            data = response.json()
        except requests.exceptions.RequestException as e:
            logging.error(f"Erreur lors de la récupération des données : {e}")
            return Response(
                {"message": "Impossible de récupérer les données depuis l'API distante."},
                status=status.HTTP_502_BAD_GATEWAY
            )
        except ValueError as e:
            logging.error(f"Erreur de parsing JSON : {e}")
            return Response(
                {"message": "Données invalides reçues depuis l'API distante."},
                status=status.HTTP_502_BAD_GATEWAY
            )
            
        months = set()
        
        # --- Étape 2 : Transaction atomique pour l'import ---
        with transaction.atomic():  # ⛔ Début du bloc atomique
            # Nettoyage des anciens imports
            TemporalMouvementDeplace.objects.all().delete()
            MouvementDeplace.objects.all().delete()
            logging.info("Anciennes données supprimées.")
            
            for mouvement in data:
                date = service.match_date_par_mois(mouvement.get('mois'))
                nom_site = str(mouvement.get('site', {}).get('nom', '')).strip()
                
                if not date:
                    logging.warning(f"Date non trouvée pour le mois : {mouvement.get('mois')}")
                    continue
                    
                site_data = SiteDeplace.objects.filter(nom_site=nom_site.strip()).first()
                if not site_data:
                    logging.warning(f"Site non trouvé : {nom_site}")
                    continue
                
                months.add(date)
                TemporalMouvementDeplace.objects.create(
                    site=site_data,
                    date_mise_a_jour=date,
                    menages=mouvement.get('menages', 0),
                    individus=mouvement.get('individus', 0),
                    pvh=0,
                )
            
            for i, m in enumerate(sorted(months, reverse=True)):
                print(f"{i+1} - {m}")
            
            variation_result = service.generate_variation_data()
            logging.info(f"Variations: {variation_result}")

            logging.info("Importation terminée.")

        return Response(service.statistiques(), status=status.HTTP_200_OK)
    except Exception as e:
        logging.error(f"Erreur lors de l'importation: {e}")
        return Response({"message": f"Erreur lors de l'importation: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def export_data(request):
    try:
        mouvements = MouvementDeplace.objects.all().order_by('date_mise_a_jour')
        logging.info(f"Nombre de mouvements trouvés: {mouvements.count()}")
        if not mouvements.exists():
            return Response(
                {"message": "Aucune donnée à exporter"}, 
                status=status.HTTP_404_NOT_FOUND
            )
            
        # Création du workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DONNEES 2023 - 2024 - 2025"
        
        entrees_individus = 0
        sortie_individus = 0

        # En-têtes
        headers = [
            "No",
            "PROVINCE", "CODE PROVINCE", "TERRITOIRE", "CODE TERRITOIRE", "ZONE SANTE", "CODE ZONE SANTE", 
            "NOM SITE", "CODE SITE", "TYPE SITE", "LONGITUDE", "LATITUDE", 
            "DATE", "TYPE MOUVEMENT", "MENAGES", "INDIVIDUS", 
            "0-4 F", "5-11 F", "12-17 F", "18-24 F", "25-59 F", "60+ F",
            "0-4 H", "5-11 H", "12-17 H", "18-24 H", "25-59 H", "60+ H",
            "PERSONNES VIVANT AVEC UN HANDICAP", "SOUS MECANISME CCCM"
        ]
        
        # Écriture des en-têtes
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
        # Écriture des données
        for index, obj in enumerate(mouvements):
            mecanisme = 1 if obj.site.sous_mecanisme else 0
            row = [
                index + 1,
                obj.site.province, obj.site.code_province, obj.site.territoire, obj.site.code_territoire, obj.site.zone_sante, 
                obj.site.code_zone_sante, obj.site.nom_site, obj.site.code_site, obj.site.type_site, obj.site.longitude, 
                obj.site.latitude, obj.date_mise_a_jour, obj.type_mouvement, obj.menages, obj.individus, 
                obj.individus_0_4_f, obj.individus_5_11_f, obj.individus_12_17_f, obj.individus_18_24_f, obj.individus_25_59_f, 
                obj.individus_60_f, obj.individus_0_4_h, obj.individus_5_11_h, obj.individus_12_17_h, obj.individus_18_24_h, 
                obj.individus_25_59_h, obj.individus_60_h, obj.pvh, mecanisme
            ]
            ws.append(row)
            
            if obj.type_mouvement == 'entree':
                entrees_individus += obj.individus
            else:
                sortie_individus += obj.individus
            
        logging.info(f"Exportation terminee. Nombre de lignes: {index + 1}")
        
        logging.info(f"Entrees individus: {entrees_individus} - Sortie individus: {sortie_individus} - Différence: {entrees_individus - sortie_individus}")
            
        # Ajustement des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Limite à 50 caractères
            ws.column_dimensions[column_letter].width = adjusted_width
            
        logging.info(f"Exportation terminee. Nombre de lignes: {index + 1}")

        # Préparation de la réponse
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="data_importable_cccm.xlsx"'
        
        wb.save(response)
        return response
    except Exception as e:
        logging.error(f"Erreur lors de l'exportation: {e}")
        return Response({"message": f"Erreur lors de l'exportation: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def export_sites_deplaces(request):
    try:
        sites = SiteDeplace.objects.all().order_by('nom_site')
        
        logging.info(f"Nombre de sites trouvés: {sites.count()}")
        if not sites.exists():
            return Response(
                {"message": "Aucune donnée à exporter"}, 
                status=status.HTTP_404_NOT_FOUND
            )
            
        # Création du workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SITES"
        
        entrees_individus = 0
        sortie_individus = 0

        # En-têtes
        headers = [
            "No",
            "PROVINCE", "CODE PROVINCE", "TERRITOIRE", "CODE TERRITOIRE", "ZONE SANTE", "CODE ZONE SANTE", 
            "NOM SITE", "CODE SITE", "TYPE SITE", "LONGITUDE", "LATITUDE",
        ]
        
        # Écriture des en-têtes
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
        # Écriture des données
        for index, obj in enumerate(sites):
            row = [
                index + 1,
                obj.province, obj.code_province, obj.territoire, obj.code_territoire, obj.zone_sante, 
                obj.code_zone_sante, obj.nom_site, obj.code_site, obj.type_site, obj.longitude, 
                obj.latitude
            ]
            ws.append(row)
            
        logging.info(f"Exportation terminee. Nombre de lignes: {index + 1}")
        
        # Ajustement des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Limite à 50 caractères
            ws.column_dimensions[column_letter].width = adjusted_width
            
        logging.info(f"Exportation terminee. Nombre de lignes: {index + 1}")

        # Préparation de la réponse
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="data_importable_cccm.xlsx"'
        
        wb.save(response)
        return response
    except Exception as e:
        logging.error(f"Erreur lors de l'exportation: {e}")
        return Response({"message": f"Erreur lors de l'exportation: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    
# Importer les sites
@api_view(['POST'])
def import_sites(request):
    try:
        site_file = request.FILES.get('site_file')
        
        if not site_file:
            return Response({"message": "Le fichier de données est requis"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(site_file, data_only=True, read_only=True)
        except Exception as e:
            return Response({"message": f"Erreur lors du chargement du fichier Excel: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        
        # Lecture et traitement des feuilles Excel
        sheet = wb['SITES']
        
        SiteDeplace.objects.all().delete()
        logging.info("Anciennes données supprimées.")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            site, _ = SiteDeplace.objects.get_or_create(
                nom_site=row[7],
                defaults={
                    'province': row[1],
                    'code_province': row[2],
                    'territoire': row[3],
                    'code_territoire': row[4],
                    'zone_sante': row[5],
                    'code_zone_sante': row[6],
                    'code_site': row[8],
                    'type_site': row[9],
                    'longitude': row[10],
                    'latitude': row[11],
                }
            )
            logging.info(f"Site '{site.nom_site}' importé avec succès.")
        
        logging.info("Données importées avec succès.")
        
        return Response({"message": "Données importées avec succès."}, status=status.HTTP_200_OK)
    except Exception as e:
        logging.error(f"Erreur lors de l'importation des données: {e}")
        return Response({"message": f"Erreur lors de l'importation des données: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def login_user(request):
    try:
        if request.user.is_authenticated:
            return redirect('mouvements_deplaces')
        
        username = request.POST.get('username').strip()
        password = request.POST.get('password').strip()
                
        if not username or not password:
            context = {
                'error_message': 'Veuillez fournir un nom d’utilisateur et un mot de passe.'
            }
            return render(request, 'login.html', context)
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('mouvements_deplaces')
        else:
            context = {
                'error_message': 'Nom d’utilisateur ou mot de passe incorrect.'
            }
            return render(request, 'login.html', context)
    except Exception as e:
        return render(request, 'login.html')
    
@api_view(['GET'])
def generate_users(request):
    try:
        users_credentials = [
            {"username": "marie", "password": "123456"},
            {"username": "paul", "password": "654321"},
            {"username": "julien", "password": "111111"},
            {"username": "amina", "password": "000000"},
            {"username": "kevin", "password": "222222"},
            {"username": "fatou", "password": "333333"},
            {"username": "luc", "password": "444444"},
            {"username": "sophie", "password": "555555"},
            {"username": "hugo", "password": "666666"},
            {"username": "ines", "password": "777777"},
            {"username": "adam", "password": "888888"},
            {"username": "nadine", "password": "999999"},
            {"username": "theo", "password": "112233"},
            {"username": "rachel", "password": "223344"},
            {"username": "yann", "password": "334455"},
            {"username": "leo", "password": "445566"},
            {"username": "claire", "password": "556677"},
            {"username": "mohamed", "password": "667788"},
            {"username": "lina", "password": "778899"},
            {"username": "nathan", "password": "889900"}
        ]
        created_users = []
        for user in users_credentials:
            user_obj, _ = User.objects.get_or_create(
                username=user["username"],
                defaults={"email": f"{user['username']}@gmail.com"},
            )
            user_obj.set_password(user["password"])
            user_obj.save()
            created_users.append({
                "username": user_obj.username,
                "password": user["password"]
            })
        return Response({
            "message": "Utilisateurs ajoutés avec succès.",
            "users": created_users
        }, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@login_required
def mouvements_deplaces(request):
    try:
        user_connect = request.user
        errors_messages = None
        if request.method == 'POST':
            mouvement_form = MouvementDeplaceSiteUniqueForm(request.POST)
            if mouvement_form.is_valid():
                f = mouvement_form.save(commit=False)
                f.user = user_connect
                f.save()
                messages.success(request, 'Mouvement deplacement ajouté avec succès.')
                return redirect('mouvements_deplaces')
            else:
                errors_messages = mouvement_form.errors
                messages.error(request, 'Formulaire invalide. Veuillez corriger les erreurs.')
        
        if user_connect.is_superuser:
            mouvements = MouvementDeplaceSiteUnique.objects.all().order_by('mois', 'site__nom')
        else:
            mouvements = MouvementDeplaceSiteUnique.objects.filter(user=user_connect).order_by('mois', 'site__nom')
        
        total_mouvements = mouvements.count()
        total_menages = sum([m.menages for m in mouvements])
        total_individus = sum([m.individus for m in mouvements])
        
        context = {
            'mouvements': mouvements,
            'mois_choices': MouvementDeplaceSiteUnique.MOIS_CHOICE,
            'sites': SiteUnique.objects.all().order_by('nom'),
            'errors_messages': errors_messages,
            'total_mouvements': total_mouvements,
            'total_menages': total_menages,
            'total_individus': total_individus
        }
        return render(request, 'mouvements_deplaces.html', context)
    except Exception as e:
        return render(request, 'mouvements_deplaces.html', {'error_message': str(e)})

@api_view(['GET'])
def mouvements_deplaces_api(request):
    try:
        mouvements = MouvementDeplaceSiteUnique.objects.all().order_by('site__nom')
        serializer = MouvementDeplaceSiteUniqueSerializer(mouvements, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return render(request, 'mouvements_deplaces.html', {'error_message': str(e)})
