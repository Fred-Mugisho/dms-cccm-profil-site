from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import *
from .serializers import *
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from collections import defaultdict
from utils.functions import *
from django.forms.models import model_to_dict

@api_view(['GET'])
def locations(request):
    try:
        provinces = Province.objects.all().order_by('name')
        serializer = ProvinceSerializer(provinces, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST', 'PUT'])
def load_locations_from_excel(request):
    try:
        locations_data = request.FILES.get('locations')
        if not locations_data:
            return Response({"message": "Fichier des localisations manquant"}, status=status.HTTP_400_BAD_REQUEST)

        wb = openpyxl.load_workbook(locations_data)
        sheets_name = ["Provinces", "Territoires", "Secteurs", "Groupements", "Villages", "Zones de sante", "Aires de sante"]
        
        for sheet_name in sheets_name:
            if sheet_name not in wb.sheetnames:
                continue  # Ignore les feuilles non présentes
            
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:  # Ignore les lignes vides
                    continue
                
                if sheet_name == "Provinces":
                    Province.objects.update_or_create(
                        code=row[0], 
                        defaults={'name': row[1]}
                    )

                elif sheet_name == "Territoires":
                    province = Province.objects.filter(code=row[2]).first()
                    if not province:
                        continue  # Ou collecter les erreurs
                    Territoire.objects.update_or_create(
                        code=row[0], 
                        defaults={'name': row[1], 'province': province}
                    )
                    
                elif sheet_name == "Secteurs":
                    territoire = Territoire.objects.filter(code=row[2]).first()
                    if not territoire:
                        continue
                    Secteur.objects.update_or_create(
                        code=row[0], 
                        defaults={'name': row[1], 'territoire': territoire}
                    )

                elif sheet_name == "Groupements":
                    secteur = Secteur.objects.filter(code=row[2]).first()
                    if not secteur:
                        continue
                    Groupement.objects.update_or_create(
                        code=row[0], 
                        defaults={'name': row[1], 'secteur': secteur}
                    )

                elif sheet_name == "Villages":
                    groupement = Groupement.objects.filter(code=row[2]).first()
                    if not groupement:
                        continue
                    Village.objects.update_or_create(
                        code=row[0], 
                        defaults={'name': row[1], 'groupement': groupement, 'latitude': row[3], 'longitude': row[4]}
                    )

                elif sheet_name == "Zones de sante":
                    territoire = Territoire.objects.filter(code=row[2]).first()
                    if not territoire:
                        continue
                    ZoneSante.objects.update_or_create(
                        code=row[0], 
                        defaults={'name': row[1], 'territoire': territoire, 'latitude': row[3], 'longitude': row[4]}
                    )
                    
                elif sheet_name == "Aires de sante":
                    health_zone = ZoneSante.objects.filter(code=row[2]).first()
                    if not health_zone:
                        continue
                    AireSante.objects.update_or_create(
                        code=row[0], 
                        defaults={'name': row[1], 'health_zone': health_zone, 'latitude': row[3], 'longitude': row[4]}
                    )

        return Response({"message": "Les données des localisations ont été chargées avec succès"}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"message": f"Erreur : {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def get_province(request, id):
    try:
        province = Province.objects.get(id=id)
        serializer = ProvinceSerializer(province)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Province.DoesNotExist:
        return Response({"message": "Province n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST', 'PUT'])
def create_update_province(request, id=None):
    try:
        province = Province.objects.get(id=id) if id else None
        province_form = ProvinceFormSerializer(province, data=request.data)
        if province_form.is_valid():
            province_form.save()
            return Response(province_form.data, status=status.HTTP_200_OK)
        else:
            return Response(province_form.errors, status=status.HTTP_400_BAD_REQUEST)
    except Province.DoesNotExist:
        return Response({"message": "Province n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def get_territoire(request, id):
    try:
        territoire = Territoire.objects.get(id=id)
        serializer = TerritoireSerializer(territoire)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Territoire.DoesNotExist:
        return Response({"message": "Territoire n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST', 'PUT'])
def create_update_territoire(request, id=None):
    try:
        territoire = Territoire.objects.get(id=id) if id else None
        territoire_form = TerritoireFormSerializer(territoire, data=request.data)
        if territoire_form.is_valid():
            territoire_form.save()
            return Response(territoire_form.data, status=status.HTTP_200_OK)
        else:
            return Response(territoire_form.errors, status=status.HTTP_400_BAD_REQUEST)
    except Territoire.DoesNotExist:
        return Response({"message": "Territoire n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def get_secteur(request, id):
    try:
        secteur = Secteur.objects.get(id=id)
        serializer = SecteurSerializer(secteur)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Secteur.DoesNotExist:
        return Response({"message": "Secteur n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST', 'PUT'])
def create_update_secteur(request, id=None):
    try:
        secteur = Secteur.objects.get(id=id) if id else None
        secteur_form = SecteurFormSerializer(secteur, data=request.data)
        if secteur_form.is_valid():
            secteur_form.save()
            return Response(secteur_form.data, status=status.HTTP_200_OK)
        else:
            return Response(secteur_form.errors, status=status.HTTP_400_BAD_REQUEST)
    except Secteur.DoesNotExist:
        return Response({"message": "Secteur n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
@api_view(['GET'])
def get_groupement(request, id):
    try:
        groupement = Groupement.objects.get(id=id)
        serializer = GroupementSerializer(groupement)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Groupement.DoesNotExist:
        return Response({"message": "Groupement n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST', 'PUT'])
def create_update_groupement(request, id=None):
    try:
        groupement = Groupement.objects.get(id=id) if id else None
        groupement_form = GroupementFormSerializer(groupement, data=request.data)
        if groupement_form.is_valid():
            groupement_form.save()
            return Response(groupement_form.data, status=status.HTTP_200_OK)
        else:
            return Response(groupement_form.errors, status=status.HTTP_400_BAD_REQUEST)
    except Groupement.DoesNotExist:
        return Response({"message": "Groupement n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def get_village(request, id):
    try:
        village = Village.objects.get(id=id)
        serializer = VillageSerializer(village)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except Village.DoesNotExist:
        return Response({"message": "Village n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST', 'PUT'])
def create_update_village(request, id=None):
    try:
        village = Village.objects.get(id=id) if id else None
        village_form = VillageFormSerializer(village, data=request.data)
        if village_form.is_valid():
            village_form.save()
            return Response(village_form.data, status=status.HTTP_200_OK)
        else:
            return Response(village_form.errors, status=status.HTTP_400_BAD_REQUEST)
    except Village.DoesNotExist:
        return Response({"message": "Village n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def get_zone_sante(request, id):
    try:
        zone_sante = ZoneSante.objects.get(id=id)
        serializer = ZoneSanteSerializer(zone_sante)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except ZoneSante.DoesNotExist:
        return Response({"message": "Zone sante n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST', 'PUT'])
def create_update_zone_sante(request, id=None):
    try:
        zone_sante = ZoneSante.objects.get(id=id) if id else None
        zone_sante_form = ZoneSanteFormSerializer(zone_sante, data=request.data)
        if zone_sante_form.is_valid():
            zone_sante_form.save()
            return Response(zone_sante_form.data, status=status.HTTP_200_OK)
        else:
            return Response(zone_sante_form.errors, status=status.HTTP_400_BAD_REQUEST)
    except ZoneSante.DoesNotExist:
        return Response({"message": "Zone sante n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['GET'])
def get_aire_sante(request, id):
    try:
        aire_sante = AireSante.objects.get(id=id)
        serializer = AireSanteSerializer(aire_sante)
        return Response(serializer.data, status=status.HTTP_200_OK)
    except AireSante.DoesNotExist:
        return Response({"message": "Aire sante n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST', 'PUT'])
def create_update_aire_sante(request, id=None):
    try:
        aire_sante = AireSante.objects.get(id=id) if id else None
        aire_sante_form = AireSanteFormSerializer(aire_sante, data=request.data)
        if aire_sante_form.is_valid():
            aire_sante_form.save()
            return Response(aire_sante_form.data, status=status.HTTP_200_OK)
        else:
            return Response(aire_sante_form.errors, status=status.HTTP_400_BAD_REQUEST)
    except AireSante.DoesNotExist:
        return Response({"message": "Aire sante n'existe pas"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)